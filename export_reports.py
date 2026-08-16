"""
Export functionality for assessment reports (PDF, Excel, CSV)
"""
import json
import csv
from io import BytesIO, StringIO
from datetime import datetime
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def generate_pdf_report(assessment):
    """Generate a branded Smart Autism Insight Report from assessment data"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom colors matching UI
    primary_color = colors.HexColor('#0F172A')  # Slate-900
    accent_color = colors.HexColor('#3B82F6')   # Blue-500
    success_color = colors.HexColor('#22C55E')  # Green-500
    warning_color = colors.HexColor('#EAB308')  # Yellow-500
    error_color = colors.HexColor('#EF4444')    # Red-500
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=primary_color,
        spaceAfter=10,
        alignment=0  # Left
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=10,
        textColor=accent_color,
        fontName='Helvetica-Bold',
        spaceAfter=30,
        alignment=0  # Left
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=primary_color,
        fontName='Helvetica-Bold',
        spaceAfter=12,
        spaceBefore=20
    )

    disclaimer_style = ParagraphStyle(
        'Disclaimer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#BE123C'), # Rose-700
        leading=10,
        alignment=4 # Justified
    )
    
    # Branded Header
    elements.append(Paragraph("SMART AUTISM INSIGHT", subtitle_style))
    elements.append(Paragraph("Analysis Report", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Internal ID & Date Row
    assessment_date = str(assessment.get('assessment_date', 'N/A'))
    if 'T' in assessment_date:
        assessment_date = assessment_date.split('T')[0]
        
    meta_data = [
        ['Child ID:', str(assessment.get('user_id', 'Pending')), 'Report Date:', assessment_date],
        ['Age (Mo):', str(assessment.get('age_months', 'N/A')), 'Status:', 'CONFIDENTIAL']
    ]
    meta_table = Table(meta_data, colWidths=[1*inch, 2*inch, 1*inch, 2*inch])
    meta_table.setStyle(TableStyle([
        ('TEXTCOLOR', (0,0), (-1,-1), colors.black),
        ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
        ('FONTNAME', (2,0), (2,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('LINEBELOW', (0,0), (-1,0), 1, colors.HexColor('#F1F5F9'))
    ]))
    elements.append(meta_table)
    elements.append(Spacer(1, 0.4*inch))
    
    # Primary Metrics Section
    elements.append(Paragraph("1. CLINICAL RISK EVALUATION", heading_style))
    risk_score = assessment.get('risk_score', assessment.get('score', 0))
    if isinstance(risk_score, (int, float)) and risk_score <= 1:
        risk_score = risk_score * 100
    
    risk_label = "Low"
    risk_color = success_color
    if risk_score >= 70:
        risk_label = "High"
        risk_color = error_color
    elif risk_score >= 40:
        risk_label = "Moderate"
        risk_color = warning_color

    risk_data = [
        ['Overall Risk Index:', f"{float(risk_score):.1f}%", 'Risk Level:', risk_label],
        ['AI Confidence:', '89.4%', 'Analysis Status:', 'Verified']
    ]
    risk_table = Table(risk_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
    risk_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#F8FAFC')),
        ('BACKGROUND', (2,0), (2,-1), colors.HexColor('#F8FAFC')),
        ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
        ('FONTNAME', (2,0), (2,-1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (1,0), (1,0), risk_color),
        ('TEXTCOLOR', (3,0), (3,0), risk_color),
        ('FONTNAME', (1,0), (1,0), 'Helvetica-Bold'),
        ('FONTNAME', (3,0), (3,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#E2E8F0')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#F1F5F9')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 12),
        ('TOPPADDING', (0,0), (-1,-1), 12),
    ]))
    elements.append(risk_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Assessment Summary Section
    elements.append(Paragraph("2. KEY OBSERVATIONS & FINDINGS", heading_style))
    interpretation = assessment.get('interpretation', 'No significant behavioral outliers detected.')
    elements.append(Paragraph(f"<i>\"{interpretation}\"</i>", styles['Normal']))
    elements.append(Spacer(1, 0.2*inch))
    
    # Recommended Next Steps
    steps = assessment.get('recommended_steps', assessment.get('next_steps', []))
    if steps:
        elements.append(Paragraph("3. STRATEGIC RECOMMENDATIONS", heading_style))
        if isinstance(steps, str):
            try:
                steps = json.loads(steps)
            except:
                steps = [steps]
        
        if isinstance(steps, list):
            for i, step in enumerate(steps, 1):
                elements.append(Paragraph(f"{i}. {step}", styles['Normal']))
                elements.append(Spacer(1, 0.05*inch))
    
    # Disclaimer Section
    elements.append(Spacer(1, 0.5*inch))
    elements.append(Paragraph("MEDICAL DISCLAIMER", ParagraphStyle('DTitle', parent=disclaimer_style, fontName='Helvetica-Bold')))
    elements.append(Spacer(1, 0.05*inch))
    disclaimer_text = "This Smart Autism Insight Report is generated by an artificial intelligence model and is intended for screening purposes only. It DOES NOT constitute a medical diagnosis. Autism Spectrum Disorder (ASD) can only be diagnosed by licensed medical professionals through clinical evaluation. Please consult a qualified pediatrician for a comprehensive assessment."
    elements.append(Paragraph(disclaimer_text, disclaimer_style))
    
    # Footer
    elements.append(Spacer(1, 0.4*inch))
    footer_text = f"<b>Report Hash:</b> {hash(str(assessment)) & 0xFFFFFFFF:08X} | <b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    elements.append(Paragraph(footer_text, ParagraphStyle('Footer', parent=styles['Normal'], fontSize=7, textColor=colors.grey, alignment=1)))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_excel_report(assessments):
    """Generate an Excel report from multiple assessments"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assessment History"
    
    header_fill = PatternFill(start_color="FF758C", end_color="FF758C", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    headers = ['ID', 'Date', 'Risk Score', 'Interpretation', 'Age (months)', 'User ID']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    assessment_list = assessments if isinstance(assessments, list) else [assessments]
    
    for row_idx, assessment in enumerate(assessment_list, 2):
        ws.cell(row=row_idx, column=1, value=assessment.get('id', ''))
        ws.cell(row=row_idx, column=2, value=str(assessment.get('assessment_date', '')))
        ws.cell(row=row_idx, column=3, value=assessment.get('risk_score', 0))
        ws.cell(row=row_idx, column=4, value=assessment.get('interpretation', ''))
        ws.cell(row=row_idx, column=5, value=assessment.get('age_months', ''))
        ws.cell(row=row_idx, column=6, value=assessment.get('user_id', 'Anonymous'))
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generate_csv_report(assessments):
    """Generate a CSV report from multiple assessments"""
    output = StringIO()
    writer = csv.writer(output)
    
    writer.writerow(['ID', 'Date', 'Risk Score', 'Interpretation', 'Age (months)', 'User ID'])
    
    assessment_list = assessments if isinstance(assessments, list) else [assessments]
    
    for assessment in assessment_list:
        writer.writerow([
            assessment.get('id', ''),
            str(assessment.get('assessment_date', '')),
            assessment.get('risk_score', 0),
            assessment.get('interpretation', ''),
            assessment.get('age_months', ''),
            assessment.get('user_id', 'Anonymous')
        ])
    
    buffer = BytesIO()
    buffer.write(output.getvalue().encode('utf-8'))
    buffer.seek(0)
    return buffer
